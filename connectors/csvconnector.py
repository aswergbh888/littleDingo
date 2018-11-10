# -*- coding: utf-8 -*-
from elasticsearch import Elasticsearch
import json
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
from itertools import chain
from connectors.connector import Connector

class Cache(object):
    def __init__(self, filepath):
        self._cache = list()
        self._filepath = filepath

    def clear(self):
        self._cache.clear()

    def append(self, data):
        self._cache.append(data)

    def size(self):
        return len(self._cache)

    def getFilepath(self):
        return self._filepath

    def getAllRows(self):
        return self._cache

class CSVConnector(Connector):

    def __init__(self, connection_info):
        Connector.__init__(self, connection_info)
        self.cache = Cache('')
        return self.connect(connection_info)

    def connect(self, info):
        return

    def startBeforeProcess(self, data):
        try:
            path = self.connection_info['host']
            shutil.rmtree(path)
        except Exception as ex:
            print(ex)

    def startProcess(self, data):
        host = self.connection_info['host']
        db = data['db']
        dir_path = os.path.join(host, db) 
        try:
            print('create dir:%s' % dir_path)
            os.makedirs(dir_path)
        except:
            pass

        filepath = os.path.join(dir_path, data['table'] + '.xlsx')
        if self.cache.getFilepath() == filepath or os.path.exists(filepath):
            pass
        else: # new table
            if self.cache.size() > 0:
                wb = load_workbook(self.cache.getFilepath())
                self.flushToCSV()

            wb = Workbook()
            sheet = wb.active
            sheet.append(list(data['row'].keys()))
            wb.save(filepath)
            self.cache = Cache(filepath)

        self.cache.append([v for v in data['row'].values()])

        if self.cache.size() >= 100:
            self.flushToCSV()

    def flushToCSV(self):
        print('==========================>flush:%s' % self.cache.getAllRows())
        filepath = self.cache.getFilepath()
        wb = load_workbook(filepath)
        sheet = wb.active
        for l in self.cache.getAllRows():
            sheet.append(l)
        wb.save(filepath)
        self.cache.clear()
